"""
Página para inserção de dados de pit stops
Permite ao estagiário inserir dados de cada etapa diretamente no aplicativo
"""

import pandas as pd
import streamlit as st
from utils.constants import (
    get_drivers_names, PITSTOP_FILE, MATTHEIS_FILE,
    calendario_2025, etapas_2025, mattheis_names_2025, mattheis_names_2024, circuit_names,
    USE_GOOGLE_SHEETS, GOOGLE_SHEETS_PITSTOP_ID, GOOGLE_SHEETS_MATTHEIS_ID,
    init_google_sheets_config
)

from utils.season_2025 import get_nome_corrida_2025, parse_corrida_name
from openpyxl import load_workbook
import os

# Configuração da página - DEVE SER O PRIMEIRO COMANDO STREAMLIT
st.set_page_config(
    page_title="Inserir Dados - Pit Stop Report",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inicializar configurações do Google Sheets (depois do set_page_config)
init_google_sheets_config()

# Re-importar variáveis atualizadas
from utils.constants import USE_GOOGLE_SHEETS, GOOGLE_SHEETS_PITSTOP_ID, GOOGLE_SHEETS_MATTHEIS_ID

# Importar Google Sheets se estiver habilitado
if USE_GOOGLE_SHEETS:
    try:
        from utils.google_sheets import save_sheet_data, load_sheet_data
    except ImportError:
        pass

st.title("📝 Inserir Dados de Pit Stops")
st.markdown("---")

# Seleção de temporada e corrida
st.subheader("1. Seleção da Corrida")

temporada = st.selectbox("Temporada:", ["2025", "2024"], index=0)

if temporada == "2025":
    # Usar calendário 2025
    corridas_disponiveis = []
    for num_corrida in range(1, 24):
        nome_corrida = get_nome_corrida_2025(num_corrida)
        corridas_disponiveis.append((num_corrida, nome_corrida))
    
    corrida_selecionada = st.selectbox(
        "Selecione a corrida:",
        options=[f"{num} - {nome}" for num, nome in corridas_disponiveis],
        format_func=lambda x: x.split(" - ", 1)[1] if " - " in x else x
    )
    
    numero_corrida = int(corrida_selecionada.split(" - ")[0])
    
    # Usar nome formatado para exibição
    st.info(f"📋 **Corrida selecionada:** {get_nome_corrida_2025(numero_corrida)}")
    info_corrida = calendario_2025[numero_corrida]
    tipo_corrida = info_corrida['tipo']
    etapa = info_corrida.get('etapa', 1)
    circuito = info_corrida.get('circuito', etapas_2025.get(etapa, 'Interlagos'))
    
    # Nome da aba no Excel (formato: E1, E1S, E2, E2S, etc.)
    if numero_corrida == 1:
        nome_aba = "E1S"  # Corrida especial
    else:
        if tipo_corrida == "Sprint":
            nome_aba = f"E{etapa}S"
        else:
            nome_aba = f"E{etapa}"
    
else:
    # Para 2024, usar formato antigo
    st.info("Para 2024, use o formato antigo de inserção.")
    nome_aba = st.text_input("Nome da aba no Excel (ex: E1, E1S):", value="E1")
    tipo_corrida = st.selectbox("Tipo de corrida:", ["Sprint", "Principal"])
    etapa = st.number_input("Etapa:", min_value=1, max_value=12, value=1)
    circuito = st.text_input("Circuito:", value="Interlagos")
    numero_corrida = None

st.markdown("---")
st.subheader("2. Dados dos Pilotos")

# Informações sobre a corrida
col1, col2, col3 = st.columns(3)
with col1:
    st.metric("Tipo de Corrida", tipo_corrida)
with col2:
    if temporada == "2025":
        st.metric("Etapa", etapa)
    else:
        st.metric("Etapa", "-")
with col3:
    if temporada == "2025":
        st.metric("Circuito", circuito)
    else:
        st.metric("Circuito", "-")

st.markdown("---")

# Modo de entrada: Individual ou em lote
modo_entrada = st.radio(
    "Modo de entrada:",
    ["Individual (um piloto por vez)", "Planilha (upload CSV/Excel)"],
    horizontal=True
)

if modo_entrada == "Individual (um piloto por vez)":
    st.subheader("3. Dados do Piloto")
    
    # Seleção do piloto - usar dicionário correto baseado na temporada
    drivers_names_season = get_drivers_names(2025 if temporada == "2025" else 2024)
    piloto_numeral = st.selectbox(
        "Numeral do Piloto:",
        options=sorted([int(k) for k in drivers_names_season.keys()]),
        format_func=lambda x: f"{x} - {drivers_names_season.get(str(x), 'Desconhecido')}"
    )
    
    # Campos de dados básicos
    col1, col2 = st.columns(2)
    with col1:
        # Posição na corrida - permite número ou "DNF"
        posicao_input = st.text_input("Posição na corrida:", value="1", help="Digite a posição numérica ou 'DNF' (Did Not Finish)")
        # Validar e converter posição
        posicao_input_upper = posicao_input.upper().strip()
        if posicao_input_upper in ["DNF", "D.N.F.", "NÃO TERMINOU", "NAO TERMINOU"]:
            posicao = "DNF"
        else:
            try:
                posicao = int(posicao_input)
                if posicao < 1:
                    st.warning("⚠️ Posição deve ser maior que 0 ou 'DNF'")
                    posicao = 1
            except ValueError:
                st.warning("⚠️ Digite um número válido ou 'DNF'")
                posicao = 1
        
        volta_pit = st.number_input("Volta do pit (pitlap):", min_value=1, value=1)
        tempo_total = st.number_input("Tempo Total (segundos):", min_value=0.0, value=0.0, step=0.01, format="%.2f")
        
        # Tempo de Troca de Pneus - permite número ou "Não registrado"
        tempo_pneu_input = st.text_input("Tempo de Troca de Pneus:", value="0.00", help="Digite o tempo em segundos ou 'Não registrado'")
        # Validar e converter tempo de pneu
        tempo_pneu_input_lower = tempo_pneu_input.lower().strip()
        if tempo_pneu_input_lower in ["não registrado", "nao registrado", "nao", "nr", "n/r"]:
            tempo_pneu = "Não registrado"
        else:
            try:
                tempo_pneu = float(tempo_pneu_input)
                if tempo_pneu < 0:
                    st.warning("⚠️ Tempo deve ser maior ou igual a 0 ou 'Não registrado'")
                    tempo_pneu = 0.0
            except ValueError:
                st.warning("⚠️ Digite um número válido ou 'Não registrado'")
                tempo_pneu = 0.0
    
    with col2:
        pneu1 = st.selectbox("Pneu 1:", ["TD", "TE", "DD", "DE", "ALL", "Não registrado"])
        if tipo_corrida == "Principal":
            pneu2 = st.selectbox("Pneu 2:", ["TD", "TE", "DD", "DE", "", "Não registrado"])
        else:
            pneu2 = ""
        
        # Campos específicos para Mattheis
        if str(piloto_numeral) in mattheis_names_2025:
            st.markdown("**Dados específicos do Grupo Mattheis:**")
            tempo_aj = st.number_input("Tempo AJ (Air Jack) - segundos:", min_value=0.0, value=0.0, step=0.01, format="%.2f")
            tempo_1c = st.number_input("Tempo 1ª Conexão (1c) - segundos:", min_value=0.0, value=0.0, step=0.01, format="%.2f")
            troca1 = st.number_input("Troca 1 (Pistola e Encaixe) - segundos:", min_value=0.0, value=0.0, step=0.01, format="%.2f")
            if tipo_corrida == "Principal":
                troca2 = st.number_input("Troca 2 (Pistola e Encaixe) - segundos:", min_value=0.0, value=0.0, step=0.01, format="%.2f")
            else:
                troca2 = 0.0
            link_youtube = st.text_input("Link do vídeo do YouTube:")
        else:
            tempo_aj = 0.0
            tempo_1c = 0.0
            troca1 = 0.0
            troca2 = 0.0
            link_youtube = ""
    
    # Botão para adicionar piloto
    if st.button("➕ Adicionar Piloto", type="primary"):
        # Criar dicionário com os dados do piloto
        # Tratar valores especiais (DNF, Não registrado)
        dados_piloto = {
            'Numeral': piloto_numeral,
            'POS': posicao,  # Pode ser número ou "DNF"
            'pitlap': volta_pit,
            'TempoTotal': tempo_total,
            'Tempopneu': tempo_pneu,  # Pode ser número ou "Não registrado"
            'Pneu1': pneu1,  # Pode ser "Não registrado"
            'Pneu2': pneu2 if pneu2 and pneu2 != "" else None,  # Pode ser "Não registrado"
            'trackid': circuit_names.get(circuito, 3) if temporada == "2025" else 1,
            'raceid': numero_corrida if numero_corrida else 1
        }
        
        # Inicializar session state para armazenar dados
        if 'dados_corrida' not in st.session_state:
            st.session_state.dados_corrida = []
        
        # Verificar se piloto já foi adicionado
        piloto_existente = next(
            (p for p in st.session_state.dados_corrida if p['Numeral'] == piloto_numeral),
            None
        )
        
        if piloto_existente:
            st.warning(f"⚠️ Piloto {piloto_numeral} já foi adicionado. Substituindo dados anteriores.")
            st.session_state.dados_corrida.remove(piloto_existente)
        
        st.session_state.dados_corrida.append(dados_piloto)
        
        # Se for piloto Mattheis, adicionar dados específicos
        mattheis_names_season = mattheis_names_2025 if temporada == "2025" else mattheis_names_2024
        if str(piloto_numeral) in mattheis_names_season:
            dados_mattheis = {
                'Numeral': piloto_numeral,
                'TempoTotal': tempo_total,
                'Tempopneu': tempo_pneu,
                'aj': tempo_aj,
                '1c': tempo_1c,
                'Troca1': troca1,
                'Troca2': troca2 if tipo_corrida == "Principal" else None,
                'link': link_youtube if link_youtube else None
            }
            
            if 'dados_mattheis' not in st.session_state:
                st.session_state.dados_mattheis = []
            
            mattheis_existente = next(
                (m for m in st.session_state.dados_mattheis if m['Numeral'] == piloto_numeral),
                None
            )
            
            if mattheis_existente:
                st.session_state.dados_mattheis.remove(mattheis_existente)
            
            st.session_state.dados_mattheis.append(dados_mattheis)
        
        drivers_names_season = get_drivers_names(2025 if temporada == "2025" else 2024)
        st.success(f"✅ Piloto {piloto_numeral} ({drivers_names_season.get(str(piloto_numeral), 'Desconhecido')}) adicionado com sucesso!")
    
    # Exibir dados coletados
    if 'dados_corrida' in st.session_state and st.session_state.dados_corrida:
        st.markdown("---")
        st.subheader("📋 Dados Coletados")
        
        df_coletados = pd.DataFrame(st.session_state.dados_corrida)
        drivers_names_season = get_drivers_names(2025 if temporada == "2025" else 2024)
        df_coletados['Piloto'] = df_coletados['Numeral'].astype(str).map(drivers_names_season)
        st.dataframe(df_coletados[['Numeral', 'Piloto', 'POS', 'pitlap', 'TempoTotal', 'Tempopneu', 'Pneu1', 'Pneu2']], 
                     use_container_width=True)
        
        # Botão para salvar
        # Re-inicializar configurações
        init_google_sheets_config()
        from utils.constants import USE_GOOGLE_SHEETS, GOOGLE_SHEETS_PITSTOP_ID, GOOGLE_SHEETS_MATTHEIS_ID
        
        if USE_GOOGLE_SHEETS:
            button_text = "💾 Salvar no Banco de Dados"
        else:
            button_text = "💾 Salvar no Excel"
        
        if st.button(button_text, type="primary"):
            try:
                # Criar DataFrame com os dados novos
                df_novo = pd.DataFrame(st.session_state.dados_corrida)
                
                if df_novo.empty:
                    st.warning("⚠️ Nenhum dado para salvar!")
                    st.stop()
                
                # Carregar dados existentes e fazer merge
                df_existente = pd.DataFrame()
                use_gs = USE_GOOGLE_SHEETS and GOOGLE_SHEETS_PITSTOP_ID
                
                if use_gs:
                    # Tentar carregar dados existentes do Google Sheets
                    try:
                        df_existente = load_sheet_data(GOOGLE_SHEETS_PITSTOP_ID, nome_aba)
                    except Exception:
                        df_existente = pd.DataFrame()
                
                if df_existente.empty:
                    # Tentar carregar do Excel local como fallback
                    try:
                        if os.path.exists(PITSTOP_FILE):
                            dados_excel = pd.read_excel(PITSTOP_FILE, sheet_name=nome_aba)
                            if not dados_excel.empty:
                                df_existente = dados_excel
                    except Exception:
                        df_existente = pd.DataFrame()
                
                # Fazer merge dos dados: manter existentes e atualizar/adicionar novos
                if not df_existente.empty:
                    # Garantir que as colunas sejam as mesmas
                    if 'Numeral' in df_existente.columns and 'Numeral' in df_novo.columns:
                        # Converter Numeral para string para comparação
                        df_existente['Numeral'] = df_existente['Numeral'].astype(str)
                        df_novo['Numeral'] = df_novo['Numeral'].astype(str)
                        
                        # Remover linhas existentes com os mesmos numerais dos novos dados
                        df_existente = df_existente[~df_existente['Numeral'].isin(df_novo['Numeral'])]
                        
                        # Combinar dados existentes (sem duplicatas) com novos dados
                        df_final = pd.concat([df_existente, df_novo], ignore_index=True)
                    else:
                        # Se não há coluna Numeral, apenas concatenar
                        df_final = pd.concat([df_existente, df_novo], ignore_index=True)
                else:
                    # Se não há dados existentes, usar apenas os novos
                    df_final = df_novo.copy()
                
                # Salvar no Google Sheets ou Excel
                if use_gs:
                    # Salvar no Google Sheets
                    try:
                        success = save_sheet_data(GOOGLE_SHEETS_PITSTOP_ID, nome_aba, df_final, clear_first=True)
                        if success:
                            st.success(f"✅ Dados salvos na aba '{nome_aba}' do Google Sheets! ({len(df_novo)} piloto(s) adicionado(s)/atualizado(s))")
                            use_gs = False  # Já salvou, não precisa salvar no Excel
                        else:
                            # Se falhou, tentar Excel local como fallback
                            raise Exception("Falha ao salvar no Google Sheets")
                    except Exception as gs_error:
                        # Se falhar ao salvar no Google Sheets, tentar Excel local como fallback
                        st.warning("⚠️ Não foi possível salvar no Google Sheets. Salvando no Excel local...")
                        use_gs = False
                
                if not use_gs:
                    # Salvar no Excel local
                    if os.path.exists(PITSTOP_FILE):
                        wb_pitstop = load_workbook(PITSTOP_FILE)
                    else:
                        from openpyxl import Workbook
                        wb_pitstop = Workbook()
                        wb_pitstop.remove(wb_pitstop.active)
                    
                    if nome_aba in wb_pitstop.sheetnames:
                        wb_pitstop.remove(wb_pitstop[nome_aba])
                    
                    ws = wb_pitstop.create_sheet(nome_aba)
                    headers = list(df_final.columns)
                    ws.append(headers)
                    
                    for _, row in df_final.iterrows():
                        ws.append([row[col] for col in headers])
                    
                    wb_pitstop.save(PITSTOP_FILE)
                    st.success(f"✅ Dados salvos na aba '{nome_aba}' do arquivo PITSTOP.xlsx! ({len(df_novo)} piloto(s) adicionado(s)/atualizado(s))")
                
                # Salvar dados Mattheis se houver
                if 'dados_mattheis' in st.session_state and st.session_state.dados_mattheis:
                    df_mattheis_novo = pd.DataFrame(st.session_state.dados_mattheis)
                    
                    if not df_mattheis_novo.empty:
                        # Carregar dados Mattheis existentes e fazer merge
                        df_mattheis_existente = pd.DataFrame()
                        
                        # Re-importar variáveis atualizadas (caso tenham mudado)
                        from utils.constants import USE_GOOGLE_SHEETS, GOOGLE_SHEETS_MATTHEIS_ID
                        
                        use_gs_mattheis = USE_GOOGLE_SHEETS and GOOGLE_SHEETS_MATTHEIS_ID
                        
                        if use_gs_mattheis:
                            # Tentar carregar dados existentes do Google Sheets
                            try:
                                df_mattheis_existente = load_sheet_data(GOOGLE_SHEETS_MATTHEIS_ID, nome_aba)
                            except Exception:
                                df_mattheis_existente = pd.DataFrame()
                        
                        if df_mattheis_existente.empty:
                            # Tentar carregar do Excel local como fallback
                            try:
                                if os.path.exists(MATTHEIS_FILE):
                                    dados_excel = pd.read_excel(MATTHEIS_FILE, sheet_name=nome_aba)
                                    if not dados_excel.empty:
                                        df_mattheis_existente = dados_excel
                            except Exception:
                                df_mattheis_existente = pd.DataFrame()
                        
                        # Fazer merge dos dados Mattheis
                        if not df_mattheis_existente.empty:
                            if 'Numeral' in df_mattheis_existente.columns and 'Numeral' in df_mattheis_novo.columns:
                                df_mattheis_existente['Numeral'] = df_mattheis_existente['Numeral'].astype(str)
                                df_mattheis_novo['Numeral'] = df_mattheis_novo['Numeral'].astype(str)
                                df_mattheis_existente = df_mattheis_existente[~df_mattheis_existente['Numeral'].isin(df_mattheis_novo['Numeral'])]
                                df_mattheis_final = pd.concat([df_mattheis_existente, df_mattheis_novo], ignore_index=True)
                            else:
                                df_mattheis_final = pd.concat([df_mattheis_existente, df_mattheis_novo], ignore_index=True)
                        else:
                            df_mattheis_final = df_mattheis_novo.copy()
                        
                        if use_gs_mattheis:
                            # Salvar no Google Sheets
                            try:
                                success = save_sheet_data(GOOGLE_SHEETS_MATTHEIS_ID, nome_aba, df_mattheis_final, clear_first=True)
                                if success:
                                    st.success(f"✅ Dados Mattheis salvos na aba '{nome_aba}' do Google Sheets! ({len(df_mattheis_novo)} piloto(s) adicionado(s)/atualizado(s))")
                                    use_gs_mattheis = False
                                else:
                                    raise Exception("Falha ao salvar dados Mattheis no Google Sheets")
                            except Exception as gs_error:
                                st.warning("⚠️ Não foi possível salvar dados Mattheis no Google Sheets. Salvando no Excel local...")
                                use_gs_mattheis = False
                        
                        if not use_gs_mattheis:
                            # Salvar no Excel local
                            if os.path.exists(MATTHEIS_FILE):
                                wb_mattheis = load_workbook(MATTHEIS_FILE)
                            else:
                                from openpyxl import Workbook
                                wb_mattheis = Workbook()
                                wb_mattheis.remove(wb_mattheis.active)
                            
                            if nome_aba in wb_mattheis.sheetnames:
                                wb_mattheis.remove(wb_mattheis[nome_aba])
                            
                            ws_mattheis = wb_mattheis.create_sheet(nome_aba)
                            headers_mattheis = list(df_mattheis_final.columns)
                            ws_mattheis.append(headers_mattheis)
                            
                            for _, row in df_mattheis_final.iterrows():
                                ws_mattheis.append([row[col] for col in headers_mattheis])
                            
                            wb_mattheis.save(MATTHEIS_FILE)
                            st.success(f"✅ Dados Mattheis salvos na aba '{nome_aba}' do arquivo Mattheis.xlsx! ({len(df_mattheis_novo)} piloto(s) adicionado(s)/atualizado(s))")
                
                # Limpar session state após salvar
                st.session_state.dados_corrida = []
                if 'dados_mattheis' in st.session_state:
                    st.session_state.dados_mattheis = []
                
                st.balloons()
                
                if not USE_GOOGLE_SHEETS:
                    st.info("💡 **Nota:** Os arquivos Excel foram salvos localmente. Para atualizar no Streamlit Cloud, você precisará fazer commit e push dos arquivos para o repositório.")
                
            except Exception as e:
                st.error(f"❌ Erro ao salvar: {str(e)}")
                st.exception(e)
        
        # Botão para limpar dados
        if st.button("🗑️ Limpar Todos os Dados"):
            st.session_state.dados_corrida = []
            if 'dados_mattheis' in st.session_state:
                st.session_state.dados_mattheis = []
            st.success("Dados limpos!")
            st.rerun()

else:  # Modo Planilha
    st.subheader("3. Upload de Planilha")
    
    arquivo_upload = st.file_uploader(
        "Faça upload de um arquivo CSV ou Excel com os dados:",
        type=['csv', 'xlsx', 'xls']
    )
    
    if arquivo_upload:
        try:
            if arquivo_upload.name.endswith('.csv'):
                df_upload = pd.read_csv(arquivo_upload)
            else:
                df_upload = pd.read_excel(arquivo_upload)
            
            st.success(f"✅ Arquivo carregado com sucesso! ({len(df_upload)} linhas)")
            st.dataframe(df_upload.head(), use_container_width=True)
            
            # Verificar colunas obrigatórias
            colunas_obrigatorias = ['Numeral', 'POS', 'pitlap', 'TempoTotal', 'Tempopneu', 'Pneu1']
            colunas_faltando = [col for col in colunas_obrigatorias if col not in df_upload.columns]
            
            if colunas_faltando:
                st.error(f"❌ Colunas obrigatórias faltando: {', '.join(colunas_faltando)}")
            else:
                st.success("✅ Todas as colunas obrigatórias estão presentes!")
                
                if USE_GOOGLE_SHEETS:
                    button_text = "💾 Salvar no Banco de Dados"
                else:
                    button_text = "💾 Salvar no Excel"
                
                if st.button(button_text, type="primary"):
                    try:
                        # Re-inicializar configurações
                        init_google_sheets_config()
                        from utils.constants import USE_GOOGLE_SHEETS, GOOGLE_SHEETS_PITSTOP_ID
                        
                        if USE_GOOGLE_SHEETS and GOOGLE_SHEETS_PITSTOP_ID:
                            # Salvar no Google Sheets
                            success = save_sheet_data(GOOGLE_SHEETS_PITSTOP_ID, nome_aba, df_upload)
                            if success:
                                st.success(f"✅ Dados salvos na aba '{nome_aba}' do Google Sheets!")
                                st.balloons()
                            else:
                                st.error("❌ Erro ao salvar no Google Sheets")
                        else:
                            # Salvar no Excel local
                            if os.path.exists(PITSTOP_FILE):
                                wb_pitstop = load_workbook(PITSTOP_FILE)
                            else:
                                from openpyxl import Workbook
                                wb_pitstop = Workbook()
                                wb_pitstop.remove(wb_pitstop.active)
                            
                            if nome_aba in wb_pitstop.sheetnames:
                                wb_pitstop.remove(wb_pitstop[nome_aba])
                            
                            ws = wb_pitstop.create_sheet(nome_aba)
                            headers = list(df_upload.columns)
                            ws.append(headers)
                            
                            for _, row in df_upload.iterrows():
                                ws.append([row[col] for col in headers])
                            
                            wb_pitstop.save(PITSTOP_FILE)
                            st.success(f"✅ Dados salvos na aba '{nome_aba}' do arquivo PITSTOP.xlsx!")
                            st.balloons()
                            st.info("💡 **Nota:** Os arquivos Excel foram salvos localmente. Para atualizar no Streamlit Cloud, você precisará fazer commit e push dos arquivos para o repositório.")
                        
                    except Exception as e:
                        st.error(f"❌ Erro ao salvar: {str(e)}")
                        st.exception(e)
        
        except Exception as e:
            st.error(f"❌ Erro ao processar arquivo: {str(e)}")

st.markdown("---")
st.info("💡 **Dica:** Após salvar os dados, recarregue a página da temporada correspondente para ver os dados atualizados.")

